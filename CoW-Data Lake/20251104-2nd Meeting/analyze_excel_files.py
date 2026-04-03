import os
import openpyxl
from pathlib import Path
import json

def analyze_excel_file(file_path):
    """Analyze an Excel file and return its structure and contents"""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        file_info = {
            'file_name': os.path.basename(file_path),
            'sheets': []
        }

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Get dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column

            # Get header row (first row)
            headers = []
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value is not None else '')

            # Get sample data (first 5 rows after header)
            sample_rows = []
            for row in range(2, min(7, max_row + 1)):
                row_data = []
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else '')
                sample_rows.append(row_data)

            sheet_info = {
                'sheet_name': sheet_name,
                'rows': max_row,
                'columns': max_col,
                'headers': headers,
                'sample_data': sample_rows
            }

            file_info['sheets'].append(sheet_info)

        workbook.close()
        return file_info

    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'error': str(e)
        }

def main():
    folder_path = r"C:\Users\bichn\Downloads\CoW-Data Lake\Data Management Lists"

    excel_files = []
    for file in Path(folder_path).glob("*.xlsx"):
        if not file.name.startswith('~'):  # Skip temp files
            excel_files.append(str(file))

    excel_files.sort()

    results = []
    for file_path in excel_files:
        print(f"Analyzing: {os.path.basename(file_path)}")
        file_info = analyze_excel_file(file_path)
        results.append(file_info)

    # Print results in a readable format
    for result in results:
        print("\n" + "="*80)
        print(f"FILE: {result['file_name']}")
        print("="*80)

        if 'error' in result:
            print(f"ERROR: {result['error']}")
            continue

        for sheet in result['sheets']:
            print(f"\nSheet: {sheet['sheet_name']}")
            print(f"  Dimensions: {sheet['rows']} rows x {sheet['columns']} columns")
            print(f"  Headers: {', '.join([h for h in sheet['headers'] if h])}")

            if sheet['sample_data']:
                print(f"  Sample data (first {len(sheet['sample_data'])} rows):")
                for i, row in enumerate(sheet['sample_data'], 1):
                    # Only print non-empty rows
                    if any(cell.strip() for cell in row):
                        print(f"    Row {i}: {' | '.join([cell[:50] for cell in row if cell])}")

if __name__ == "__main__":
    main()

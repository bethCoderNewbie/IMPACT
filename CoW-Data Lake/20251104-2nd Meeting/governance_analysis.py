import os
import re
import sys
import datetime
import openpyxl
from pathlib import Path
from collections import defaultdict

def deep_analyze_excel_files(folder_path):
    """Deep analysis to extract governance-related information"""

    governance_data = {
        'data_classification': set(),
        'security_tags': set(),
        'cutoff_types': {},
        'retention_periods': set(),
        'disposition_types': set(),
        'templates': set(),
        'departments': set(),
        'document_classes': set(),
        'abbreviations': {},
        'metadata_fields': set(),
        'notes_and_requirements': [],
        'divisions_boards': {},
        'rm_series': set()
    }

    excel_files = [str(f) for f in Path(folder_path).glob("*.xlsx") if not f.name.startswith('~')]
    excel_files.sort()

    for file_path in excel_files:
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            file_name = os.path.basename(file_path)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Get headers
                headers = []
                for col in range(1, sheet.max_column + 1):
                    header = sheet.cell(row=1, column=col).value
                    if header:
                        headers.append(str(header).strip())
                        governance_data['metadata_fields'].add(str(header).strip())

                # Extract data based on headers
                for row in range(2, sheet.max_row + 1):
                    row_data = {}
                    for col, header in enumerate(headers, 1):
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value:
                            row_data[header] = str(cell_value).strip()

                    # Extract specific governance elements
                    if 'SecTag' in row_data:
                        governance_data['security_tags'].add(row_data['SecTag'])

                    if 'RMCutoff' in row_data or 'RMCuttoff' in row_data or 'Cutoff' in row_data:
                        cutoff = row_data.get('RMCutoff') or row_data.get('RMCuttoff') or row_data.get('Cutoff')
                        if cutoff and cutoff.strip():
                            governance_data['cutoff_types'][cutoff] = None

                    if 'RMRetention' in row_data or 'Retention' in row_data:
                        retention = row_data.get('RMRetention') or row_data.get('Retention')
                        if retention and retention.strip():
                            governance_data['retention_periods'].add(retention)

                    if 'Disposition' in row_data:
                        disp = row_data['Disposition']
                        if disp and disp.strip():
                            governance_data['disposition_types'].add(disp)

                    if 'Template' in row_data:
                        template = row_data['Template']
                        if template and template.strip():
                            governance_data['templates'].add(template)
                            governance_data['departments'].add(template)

                    if 'DocClass' in row_data:
                        doc_class = row_data['DocClass']
                        if doc_class and doc_class.strip():
                            governance_data['document_classes'].add(doc_class)

                    if 'Abbr' in row_data:
                        abbr = row_data['Abbr']
                        if abbr and abbr.strip() and 'DocType' in row_data:
                            governance_data['abbreviations'][abbr] = row_data['DocType']

                    if 'RMSeries' in row_data:
                        rm_series = row_data['RMSeries']
                        if rm_series and rm_series.strip():
                            governance_data['rm_series'].add(rm_series)

                    # Capture descriptions from special sheets
                    if sheet_name == 'Cutoff' and 'Description' in row_data and 'Name' in row_data:
                        governance_data['cutoff_types'][row_data['Name']] = row_data['Description']

                    # Capture department codes
                    if 'Code' in row_data and 'Short Desc' in row_data and 'Long Desc' in row_data:
                        code = row_data['Code']
                        short = row_data['Short Desc']
                        long_desc = row_data['Long Desc']
                        governance_data['abbreviations'][short] = long_desc

                    # Capture divisions/boards/districts
                    if sheet_name == 'Notes' or 'Division' in headers or 'Board' in headers or 'District' in headers:
                        for header in headers:
                            if header in ['Division', 'Board', 'District', 'Type', 'Awards', 'subfolders']:
                                if header in row_data:
                                    if file_name not in governance_data['divisions_boards']:
                                        governance_data['divisions_boards'][file_name] = []
                                    governance_data['divisions_boards'][file_name].append({
                                        'type': header,
                                        'value': row_data[header]
                                    })

                # Capture notes
                if sheet_name == 'Notes':
                    for row in range(1, sheet.max_row + 1):
                        for col in range(1, sheet.max_column + 1):
                            cell_value = sheet.cell(row=row, column=col).value
                            if cell_value and str(cell_value).strip():
                                note_text = str(cell_value).strip()
                                if len(note_text) > 5:  # Filter out very short entries
                                    governance_data['notes_and_requirements'].append({
                                        'file': file_name,
                                        'note': note_text
                                    })

            workbook.close()

        except Exception as e:
            print(f"Error processing {file_path}: {str(e)}")

    return governance_data

def print_governance_report(data):
    """Print detailed governance analysis report"""

    # Set UTF-8 encoding for console output
    import sys
    if sys.stdout.encoding != 'utf-8':
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    print("=" * 100)
    print("LASERFICHE GOVERNANCE ANALYSIS - INFORMATION AVAILABLE vs NEEDED")
    print("=" * 100)

    print("\n" + "=" * 100)
    print("1. DATA CLASSIFICATION SCHEMA")
    print("=" * 100)
    print("\nAVAILABLE IN FILES:")
    print(f"  - Security Tags found: {len(data['security_tags'])}")
    if data['security_tags']:
        for tag in sorted(data['security_tags']):
            print(f"    • {tag}")
    else:
        print("    • No explicit security tags found in the data")

    print("\nNEEDED INFORMATION:")
    print("  ❌ Data classification levels (Public, Internal, Confidential, Restricted)")
    print("  ❌ Specific definitions for each classification level")
    print("  ❌ Handling rules for each classification level")
    print("  ❌ Who can access each classification level")
    print("  ❌ Storage requirements per classification")

    print("\n" + "=" * 100)
    print("2. FILE NAMING & METADATA STANDARDS")
    print("=" * 100)
    print("\nAVAILABLE IN FILES:")
    print(f"  - Metadata Fields identified: {len(data['metadata_fields'])}")
    for field in sorted(data['metadata_fields']):
        print(f"    • {field}")

    print(f"\n  - Document Abbreviations: {len(data['abbreviations'])}")
    for abbr, full_name in sorted(list(data['abbreviations'].items())[:10]):
        print(f"    • {abbr} = {full_name}")
    if len(data['abbreviations']) > 10:
        print(f"    ... and {len(data['abbreviations']) - 10} more")

    print("\nNEEDED INFORMATION:")
    print("  ❌ Complete file naming convention schema with examples")
    print("  ❌ Naming logic/rules (e.g., [Dept]-[Type]-[Date]-[Description].ext)")
    print("  ❌ Which metadata fields are MANDATORY vs OPTIONAL")
    print("  ❌ Validation rules for metadata (formats, allowed values)")
    print("  ❌ Default values for metadata fields")
    print("  ❌ Metadata inheritance rules (folder-level vs document-level)")

    print("\n" + "=" * 100)
    print("3. RECORDS COORDINATORS / DATA STEWARDS")
    print("=" * 100)
    print("\nAVAILABLE IN FILES:")
    print("  - Partial user access requirements found in Notes sections:")
    for note in data['notes_and_requirements'][:15]:
        if 'license' in note['note'].lower() or 'access' in note['note'].lower() or 'user' in note['note'].lower():
            print(f"    • [{note['file']}] {note['note'][:80]}")

    print("\nNEEDED INFORMATION:")
    print("  ❌ Complete list of 30 records coordinators with names and contact info")
    print("  ❌ Department assignment for each coordinator")
    print("  ❌ Specific responsibilities for each coordinator")
    print("  ❌ Training requirements for coordinators")
    print("  ❌ Escalation procedures when coordinator is unavailable")
    print("  ❌ Performance metrics/KPIs for coordinators")

    print("\n" + "=" * 100)
    print("4. STAKEHOLDER ROLES (RACI MATRIX)")
    print("=" * 100)
    print("\nAVAILABLE IN FILES:")
    print("  - Departments identified: " + ", ".join(sorted(list(data['departments'])[:10])))
    if len(data['departments']) > 10:
        print(f"    ... and {len(data['departments']) - 10} more")

    print("\nNEEDED INFORMATION:")
    print("  ❌ Complete RACI matrix (Responsible, Accountable, Consulted, Informed)")
    print("  ❌ IT Department roles and responsibilities")
    print("  ❌ Legal Department roles and responsibilities")
    print("  ❌ Department Head responsibilities")
    print("  ❌ City Clerk's role in records management")
    print("  ❌ Audit and compliance oversight roles")
    print("  ❌ Escalation paths for governance issues")

    print("\n" + "=" * 100)
    print("5. RETENTION SCHEDULES")
    print("=" * 100)
    print("\nAVAILABLE IN FILES:")
    print(f"  - Cutoff Types: {len(data['cutoff_types'])}")
    for cutoff, desc in sorted(data['cutoff_types'].items()):
        if desc:
            print(f"    • {cutoff}: {desc}")
        else:
            print(f"    • {cutoff}")

    print(f"\n  - Retention Periods: {len(data['retention_periods'])}")
    for period in sorted(data['retention_periods']):
        print(f"    • {period}")

    print(f"\n  - Disposition Types:")
    for disp in sorted(data['disposition_types']):
        print(f"    • {disp}")

    print(f"\n  - Record Management Series:")
    for series in sorted(data['rm_series']):
        print(f"    • {series}")

    print("\nNEEDED INFORMATION:")
    print("  ❌ Source of retention schedules (City Clerk's office? State law? Federal?)")
    print("  ❌ How retention schedules are updated/maintained")
    print("  ❌ Legal citations for retention requirements")
    print("  ❌ Exceptions to standard retention rules")
    print("  ❌ Hold/litigation hold procedures")
    print("  ❌ How to handle documents with multiple retention requirements")

    print("\n" + "=" * 100)
    print("6. RETENTION ENFORCEMENT")
    print("=" * 100)
    print("\nAVAILABLE IN FILES:")
    print("  - Cutoff triggers defined (File Date, Expiration, Superseded, etc.)")

    print("\nNEEDED INFORMATION:")
    print("  ❌ Is retention enforcement automated or manual?")
    print("  ❌ What metadata field triggers the retention countdown?")
    print("  ❌ Notification system before disposition")
    print("  ❌ Approval workflow for disposition")
    print("  ❌ Audit trail for disposed records")
    print("  ❌ Recovery procedures for accidentally disposed records")
    print("  ❌ System configuration for automated retention")

    print("\n" + "=" * 100)
    print("7. ACCESS CONTROL MODEL")
    print("=" * 100)
    print("\nAVAILABLE IN FILES:")
    print("  - Security Tags referenced in documents")
    print("  - Departments/templates suggest departmental boundaries")

    print("\nNEEDED INFORMATION:")
    print("  ❌ Complete access control model documentation")
    print("  ❌ User role definitions (e.g., Admin, Power User, Read-Only)")
    print("  ❌ Permission levels (Read, Write, Delete, Share, etc.)")
    print("  ❌ How departmental boundaries enforce access")
    print("  ❌ Cross-department access procedures")
    print("  ❌ Privileged access management (PAM) for sensitive data")
    print("  ❌ Service account management")

    print("\n" + "=" * 100)
    print("8. SENSITIVE DATA ACCESS WORKFLOW")
    print("=" * 100)
    print("\nAVAILABLE IN FILES:")
    print("  - No explicit workflow documentation found")

    print("\nNEEDED INFORMATION:")
    print("  ❌ Definition of 'sensitive data' in this context")
    print("  ❌ Request submission process")
    print("  ❌ Approval authority matrix")
    print("  ❌ Timeframe for access request processing")
    print("  ❌ Temporary vs permanent access procedures")
    print("  ❌ Access review and recertification schedule")
    print("  ❌ Logging and monitoring of sensitive data access")

    print("\n" + "=" * 100)
    print("9. DATA EXPORT & SYSTEM CONNECTION")
    print("=" * 100)
    print("\nAVAILABLE IN FILES:")
    print("  - No export or integration documentation found")

    print("\nNEEDED INFORMATION:")
    print("  ❌ Export capabilities and restrictions")
    print("  ❌ API availability for system integration")
    print("  ❌ Data format standards for exports")
    print("  ❌ Governance rules that apply to exported data")
    print("  ❌ Metadata preservation in exports")
    print("  ❌ Integration with future data lake or analytics platforms")
    print("  ❌ Export approval workflow")
    print("  ❌ Audit trail for data exports")

    print("\n" + "=" * 100)
    print("10. GOVERNANCE ADOPTION CHALLENGES")
    print("=" * 100)
    print("\nAVAILABLE IN FILES:")
    print("  - Notes sections show requirements being gathered")
    print("  - Multiple departments with different needs")

    print("\nNEEDED INFORMATION:")
    print("  ❌ Historical adoption challenges (lessons learned)")
    print("  ❌ Change management strategy")
    print("  ❌ Training programs and materials")
    print("  ❌ Compliance monitoring and enforcement")
    print("  ❌ User feedback mechanisms")
    print("  ❌ Success metrics and adoption rates")
    print("  ❌ Ongoing governance improvements")

    print("\n" + "=" * 100)
    print("SUMMARY OF ADDITIONAL DOCUMENTS NEEDED")
    print("=" * 100)
    print("\nTo fully answer the governance questions, you need:")
    print("\n1. 📄 Data Classification Policy Document")
    print("   - Classification levels and definitions")
    print("   - Handling requirements per level")
    print("   - Access requirements per level")

    print("\n2. 📄 Laserfiche Configuration/Setup Documentation")
    print("   - File naming convention standards")
    print("   - Mandatory vs optional metadata fields")
    print("   - Template configurations")
    print("   - Validation rules")

    print("\n3. 📄 Records Coordinator List & Responsibilities")
    print("   - Names and contact information")
    print("   - Department assignments")
    print("   - Role descriptions")

    print("\n4. 📄 RACI Matrix / Governance Charter")
    print("   - Roles and responsibilities for all stakeholders")
    print("   - Decision-making authority")
    print("   - Escalation procedures")

    print("\n5. 📄 Official Retention Schedule Documentation")
    print("   - Source and legal basis")
    print("   - Update procedures")
    print("   - Exception handling")

    print("\n6. 📄 Retention Enforcement Procedures")
    print("   - Automation configuration")
    print("   - Manual procedures")
    print("   - Approval workflows")

    print("\n7. 📄 Access Control Policy & Procedures")
    print("   - Role definitions")
    print("   - Permission matrices")
    print("   - Access request workflows")

    print("\n8. 📄 Sensitive Data Handling Policy")
    print("   - Definition of sensitive data")
    print("   - Special access procedures")
    print("   - Audit requirements")

    print("\n9. 📄 Data Integration & Export Policy")
    print("   - Export procedures")
    print("   - API documentation")
    print("   - Integration governance")

    print("\n10. 📄 Change Management & Training Documentation")
    print("   - Adoption strategy")
    print("   - Training materials")
    print("   - Lessons learned")

    print("\n" + "=" * 100)
    print("KEY STAKEHOLDERS TO INTERVIEW")
    print("=" * 100)
    print("\n1. 👤 Laserfiche System Administrator(s)")
    print("2. 👤 City Clerk or Records Manager")
    print("3. 👤 IT Department Leadership")
    print("4. 👤 Legal Department Representative")
    print("5. 👤 Sample of Records Coordinators (3-5 from different departments)")
    print("6. 👤 Department Heads (especially heavy users)")
    print("7. 👤 Information Security Officer")
    print("8. 👤 Compliance/Audit Representative")

    print("\n" + "=" * 100)

def validate_file_naming_conventions(folder_path):
    """Validate Excel files follow convention: DEPT-D-###-Description.xlsx

    Expected pattern based on Data Management Lists naming:
      [2-letter dept code]-D-[3-digit seq]-[Description].xlsx
    Validates against GOVERNANCE_GAP_ANALYSIS.md §2 File Naming & Metadata Standards.
    """
    naming_pattern = re.compile(r'^([A-Z]{2})-D-(\d{3})-(.+)\.xlsx$')
    valid, invalid = [], []

    for f in sorted(Path(folder_path).glob("*.xlsx")):
        if f.name.startswith('~'):
            continue
        match = naming_pattern.match(f.name)
        if match:
            valid.append({
                'file': f.name,
                'dept_code': match.group(1),
                'seq': match.group(2),
                'description': match.group(3)
            })
        else:
            invalid.append(f.name)

    return valid, invalid


def validate_dept_codes(folder_path, valid_files):
    """Cross-reference dept codes in filenames against GE-D-001 master list.

    GE-D-001-Department Abbreviations CODES.xlsx is the authoritative source
    for valid two-letter department codes used in the naming convention.
    """
    dept_codes_file = os.path.join(folder_path, 'GE-D-001-Department Abbreviations CODES.xlsx')
    known_codes = set()

    if os.path.exists(dept_codes_file):
        try:
            wb = openpyxl.load_workbook(dept_codes_file, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                headers = []
                for col in range(1, sheet.max_column + 1):
                    h = sheet.cell(row=1, column=col).value
                    headers.append(str(h).strip() if h else '')
                for row in range(2, sheet.max_row + 1):
                    row_data = {}
                    for col, h in enumerate(headers, 1):
                        val = sheet.cell(row=row, column=col).value
                        if val and h:
                            row_data[h] = str(val).strip()
                    # GE-D-001 uses 'Code' column for dept abbreviations
                    if 'Code' in row_data and row_data['Code']:
                        known_codes.add(row_data['Code'])
            wb.close()
        except Exception as e:
            print(f"Error reading dept codes file: {e}")
    else:
        print(f"⚠️  GE-D-001 not found at: {dept_codes_file}")

    # GE is the general/global file itself — exclude from unknown check
    unknown_codes = [
        {'file': f['file'], 'dept_code': f['dept_code']}
        for f in valid_files
        if f['dept_code'] not in known_codes and f['dept_code'] != 'GE'
    ]

    return known_codes, unknown_codes


def validate_data_quality(folder_path):
    """Detect data quality issues identified in GOVERNANCE_GAP_ANALYSIS.md §Gaps/Concerns:

    - Misspelled column headers (RMCuttoff vs RMCutoff)
    - Disposition value inconsistencies: 'destroy'/'Destroy'/'Desroy'
    - Retention period case inconsistency: 'Permanent' vs 'permanent'
    - Retention fields that contain a filename instead of a period value
    """
    issues = {
        'misspelled_headers': [],
        'disposition_issues': [],
        'retention_case_issues': [],
        'retention_contains_filename': [],
    }

    DISPOSITION_CANONICAL = {'destroy', 'permanent', 'retain'}
    RETENTION_FILENAME_PATTERN = re.compile(r'[A-Z]{2}-D-\d{3}|\.xlsx', re.IGNORECASE)

    excel_files = sorted(
        str(f) for f in Path(folder_path).glob("*.xlsx") if not f.name.startswith('~')
    )

    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                headers = []
                for col in range(1, sheet.max_column + 1):
                    h = sheet.cell(row=1, column=col).value
                    headers.append(str(h).strip() if h else '')

                # --- Header-level checks ---
                if 'RMCuttoff' in headers:
                    issues['misspelled_headers'].append({
                        'file': file_name,
                        'sheet': sheet_name,
                        'found': 'RMCuttoff',
                        'expected': 'RMCutoff'
                    })

                # --- Row-level checks ---
                for row in range(2, sheet.max_row + 1):
                    row_data = {}
                    for col, h in enumerate(headers, 1):
                        val = sheet.cell(row=row, column=col).value
                        if val and h:
                            row_data[h] = str(val).strip()

                    # Disposition: case & typo check
                    if 'Disposition' in row_data:
                        disp = row_data['Disposition']
                        disp_lower = disp.lower()
                        # Typo: "Desroy"
                        if 'desroy' in disp_lower:
                            issues['disposition_issues'].append({
                                'file': file_name, 'sheet': sheet_name, 'row': row,
                                'value': disp, 'issue': "typo 'Desroy' (expected 'destroy')"
                            })
                        # Case inconsistency for known values
                        elif disp_lower in DISPOSITION_CANONICAL and disp != disp_lower:
                            issues['disposition_issues'].append({
                                'file': file_name, 'sheet': sheet_name, 'row': row,
                                'value': disp, 'issue': f"case inconsistency (expected '{disp_lower}')"
                            })

                    # Retention: case check ('Permanent' vs 'permanent')
                    retention_key = 'RMRetention' if 'RMRetention' in row_data else (
                        'Retention' if 'Retention' in row_data else None
                    )
                    if retention_key:
                        ret = row_data[retention_key]
                        if ret and ret[0].isupper() and ret.lower() in {'permanent'}:
                            issues['retention_case_issues'].append({
                                'file': file_name, 'sheet': sheet_name, 'row': row,
                                'value': ret, 'issue': f"case inconsistency (expected '{ret.lower()}')"
                            })
                        # Filename accidentally placed in retention field
                        if ret and RETENTION_FILENAME_PATTERN.search(ret):
                            issues['retention_contains_filename'].append({
                                'file': file_name, 'sheet': sheet_name, 'row': row,
                                'value': ret[:80]
                            })

            workbook.close()
        except Exception as e:
            print(f"Error validating {file_path}: {e}")

    return issues


def print_validation_report(valid_files, invalid_files, quality_issues, known_codes, unknown_codes):
    """Print naming convention & data quality validation results."""
    print("\n" + "=" * 100)
    print("VALIDATION REPORT: Naming Conventions & Data Quality")
    print("(Cross-referenced against GOVERNANCE_GAP_ANALYSIS.md)")
    print("=" * 100)

    # --- File naming ---
    print(f"\n{'─'*60}")
    print(f"📁  FILE NAMING CONVENTION  (expected: DEPT-D-###-Description.xlsx)")
    print(f"{'─'*60}")
    print(f"  ✅ Conforming files : {len(valid_files)}")
    print(f"  ❌ Non-conforming   : {len(invalid_files)}")
    for f in invalid_files:
        print(f"       • {f}")

    # --- Dept codes ---
    print(f"\n{'─'*60}")
    print(f"🏢  DEPARTMENT CODE VALIDATION  (source: GE-D-001)")
    print(f"{'─'*60}")
    if known_codes:
        print(f"  Known codes ({len(known_codes)}): {', '.join(sorted(known_codes))}")
    else:
        print("  ⚠️  Could not load GE-D-001 — dept code validation skipped")
    if unknown_codes:
        print(f"  ❌ Files using unrecognised dept codes:")
        for u in unknown_codes:
            print(f"       • {u['file']}  (code: '{u['dept_code']}')")
    else:
        print("  ✅ All dept codes match GE-D-001")

    # --- Data quality ---
    print(f"\n{'─'*60}")
    print(f"🔍  DATA QUALITY ISSUES  (per GOVERNANCE_GAP_ANALYSIS.md §Gaps/Concerns)")
    print(f"{'─'*60}")

    mh = quality_issues['misspelled_headers']
    print(f"\n  Misspelled column headers: {len(mh)}")
    for i in mh:
        print(f"    ⚠️  [{i['file']} / {i['sheet']}]  '{i['found']}'  →  should be  '{i['expected']}'")

    di = quality_issues['disposition_issues']
    print(f"\n  Disposition value issues (case / typos): {len(di)}")
    for i in di[:20]:
        print(f"    ⚠️  [{i['file']} / {i['sheet']}] row {i['row']:>4}  '{i['value']}'  →  {i['issue']}")
    if len(di) > 20:
        print(f"       … and {len(di) - 20} more")

    rc = quality_issues['retention_case_issues']
    print(f"\n  Retention case inconsistencies: {len(rc)}")
    for i in rc[:20]:
        print(f"    ⚠️  [{i['file']} / {i['sheet']}] row {i['row']:>4}  '{i['value']}'  →  {i['issue']}")
    if len(rc) > 20:
        print(f"       … and {len(rc) - 20} more")

    rf = quality_issues['retention_contains_filename']
    print(f"\n  Retention fields containing filename (data entry error): {len(rf)}")
    for i in rf:
        print(f"    ❌  [{i['file']} / {i['sheet']}] row {i['row']:>4}  '{i['value']}'")

    total = len(mh) + len(di) + len(rc) + len(rf) + len(invalid_files) + len(unknown_codes)
    print(f"\n{'─'*60}")
    print(f"  TOTAL ISSUES FOUND: {total}")
    print("=" * 100)


def get_report_path(output_dir):
    """Return a versioned report path: GOVERNANCE-REPORT_YYYY-MM-DD_vN.md

    Version N auto-increments by counting existing reports for today's date
    in output_dir, so each run produces a distinct file without overwriting.
    """
    today = datetime.date.today().strftime("%Y-%m-%d")
    prefix = f"GOVERNANCE-REPORT_{today}_v"
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    existing = sorted(output_dir.glob(f"{prefix}*.md"))
    next_version = len(existing) + 1
    return output_dir / f"{prefix}{next_version}.md"


def build_md_report(data, valid_files, invalid_files, quality_issues,
                    known_codes, unknown_codes, run_ts):
    """Build the full governance + validation report as a Markdown string."""
    lines = []
    L = lines.append

    L(f"# Laserfiche Governance Analysis Report")
    L(f"**Generated:** {run_ts}  ")
    L(f"**Source:** Data Management Lists ({len(list(Path(r'C:\\Users\\bichn\\MSBA\\IMPACT\\CoW-Data Lake\\Data Management Lists').glob('*.xlsx')))} Excel files)  ")
    L(f"**Script:** governance_analysis.py")
    L("")

    # ── PART 1: GOVERNANCE ANALYSIS ──────────────────────────────────────────
    L("---")
    L("## Part 1 — Governance Analysis: Information Available vs Needed")
    L("")

    # 1. Data Classification
    L("### 1. Data Classification Schema")
    L("")
    L("**Available in files:**")
    L(f"- Security Tags found: **{len(data['security_tags'])}**")
    if data['security_tags']:
        for tag in sorted(data['security_tags']):
            L(f"  - `{tag}`")
    else:
        L("  - No explicit security tags found in the data")
    L("")
    L("**Needed (gaps):**")
    for gap in [
        "Data classification levels (Public, Internal, Confidential, Restricted)",
        "Specific definitions for each classification level",
        "Handling rules for each classification level",
        "Who can access each classification level",
        "Storage requirements per classification",
    ]:
        L(f"- ❌ {gap}")
    L("")

    # 2. File Naming & Metadata
    L("### 2. File Naming & Metadata Standards")
    L("")
    L("**Available in files:**")
    L(f"- Metadata fields identified: **{len(data['metadata_fields'])}**")
    for field in sorted(data['metadata_fields']):
        L(f"  - `{field}`")
    L(f"- Document abbreviations: **{len(data['abbreviations'])}**")
    for abbr, full_name in sorted(list(data['abbreviations'].items())[:10]):
        L(f"  - `{abbr}` = {full_name}")
    if len(data['abbreviations']) > 10:
        L(f"  - *… and {len(data['abbreviations']) - 10} more*")
    L("")
    L("**Needed (gaps):**")
    for gap in [
        "Complete file naming convention schema with examples",
        "Naming logic/rules (e.g., [Dept]-[Type]-[Date]-[Description].ext)",
        "Which metadata fields are MANDATORY vs OPTIONAL",
        "Validation rules for metadata (formats, allowed values)",
        "Default values for metadata fields",
        "Metadata inheritance rules (folder-level vs document-level)",
    ]:
        L(f"- ❌ {gap}")
    L("")

    # 3. Records Coordinators
    L("### 3. Records Coordinators / Data Stewards")
    L("")
    L("**Available in files:**")
    L("- Partial user access requirements found in Notes sections:")
    for note in data['notes_and_requirements'][:15]:
        if any(kw in note['note'].lower() for kw in ('license', 'access', 'user')):
            L(f"  - `[{note['file']}]` {note['note'][:100]}")
    L("")
    L("**Needed (gaps):**")
    for gap in [
        "Complete list of 30 records coordinators with names and contact info",
        "Department assignment for each coordinator",
        "Specific responsibilities for each coordinator",
        "Training requirements for coordinators",
        "Escalation procedures when coordinator is unavailable",
        "Performance metrics/KPIs for coordinators",
    ]:
        L(f"- ❌ {gap}")
    L("")

    # 4. Stakeholder Roles
    L("### 4. Stakeholder Roles (RACI Matrix)")
    L("")
    L("**Available in files:**")
    dept_list = sorted(data['departments'])
    L(f"- Departments identified: {', '.join(dept_list[:10])}")
    if len(dept_list) > 10:
        L(f"  - *… and {len(dept_list) - 10} more*")
    L("")
    L("**Needed (gaps):**")
    for gap in [
        "Complete RACI matrix (Responsible, Accountable, Consulted, Informed)",
        "IT Department roles and responsibilities",
        "Legal Department roles and responsibilities",
        "Department Head responsibilities",
        "City Clerk's role in records management",
        "Audit and compliance oversight roles",
        "Escalation paths for governance issues",
    ]:
        L(f"- ❌ {gap}")
    L("")

    # 5. Retention Schedules
    L("### 5. Retention Schedules")
    L("")
    L("**Available in files:**")
    L(f"- Cutoff types: **{len(data['cutoff_types'])}**")
    for cutoff, desc in sorted(data['cutoff_types'].items()):
        L(f"  - `{cutoff}`" + (f": {desc}" if desc else ""))
    L(f"- Retention periods: **{len(data['retention_periods'])}**")
    for period in sorted(data['retention_periods']):
        L(f"  - `{period}`")
    L(f"- Disposition types:")
    for disp in sorted(data['disposition_types']):
        L(f"  - `{disp}`")
    L(f"- RM Series:")
    for series in sorted(data['rm_series']):
        L(f"  - `{series}`")
    L("")
    L("**Needed (gaps):**")
    for gap in [
        "Source of retention schedules (City Clerk's office? State law? Federal?)",
        "How retention schedules are updated/maintained",
        "Legal citations for retention requirements",
        "Exceptions to standard retention rules",
        "Hold/litigation hold procedures",
        "How to handle documents with multiple retention requirements",
    ]:
        L(f"- ❌ {gap}")
    L("")

    # 6–10 (condensed remaining sections)
    remaining = [
        ("6. Retention Enforcement", [
            "Is retention enforcement automated or manual?",
            "What metadata field triggers the retention countdown?",
            "Notification system before disposition",
            "Approval workflow for disposition",
            "Audit trail for disposed records",
            "Recovery procedures for accidentally disposed records",
            "System configuration for automated retention",
        ], "Cutoff triggers defined (File Date, Expiration, Superseded, etc.)"),
        ("7. Access Control Model", [
            "Complete access control model documentation",
            "User role definitions (e.g., Admin, Power User, Read-Only)",
            "Permission levels (Read, Write, Delete, Share, etc.)",
            "How departmental boundaries enforce access",
            "Cross-department access procedures",
            "Privileged access management (PAM) for sensitive data",
            "Service account management",
        ], "Security tags referenced; departmental boundaries implied by templates"),
        ("8. Sensitive Data Access Workflow", [
            "Definition of 'sensitive data' in this context",
            "Request submission process",
            "Approval authority matrix",
            "Timeframe for access request processing",
            "Temporary vs permanent access procedures",
            "Access review and recertification schedule",
            "Logging and monitoring of sensitive data access",
        ], "No explicit workflow documentation found"),
        ("9. Data Export & System Connection", [
            "Export capabilities and restrictions",
            "API availability for system integration",
            "Data format standards for exports",
            "Governance rules that apply to exported data",
            "Metadata preservation in exports",
            "Integration with future data lake or analytics platforms",
            "Export approval workflow",
            "Audit trail for data exports",
        ], "No export or integration documentation found"),
        ("10. Governance Adoption Challenges", [
            "Historical adoption challenges (lessons learned)",
            "Change management strategy",
            "Training programs and materials",
            "Compliance monitoring and enforcement",
            "User feedback mechanisms",
            "Success metrics and adoption rates",
            "Ongoing governance improvements",
        ], "Notes sections show requirements being gathered; 22 departments with different needs"),
    ]
    for title, gaps, available in remaining:
        L(f"### {title}")
        L("")
        L(f"**Available in files:** {available}")
        L("")
        L("**Needed (gaps):**")
        for gap in gaps:
            L(f"- ❌ {gap}")
        L("")

    # ── PART 2: VALIDATION REPORT ─────────────────────────────────────────────
    L("---")
    L("## Part 2 — Validation Report: Naming Conventions & Data Quality")
    L("")
    L("> Cross-referenced against `GOVERNANCE_GAP_ANALYSIS.md`")
    L("")

    # File naming
    L("### File Naming Convention")
    L(f"Expected pattern: `DEPT-D-###-Description.xlsx`")
    L("")
    L(f"| Result | Count |")
    L(f"|--------|-------|")
    L(f"| ✅ Conforming | {len(valid_files)} |")
    L(f"| ❌ Non-conforming | {len(invalid_files)} |")
    if invalid_files:
        L("")
        L("**Non-conforming files:**")
        for f in invalid_files:
            L(f"- `{f}`")
    L("")

    # Dept codes
    L("### Department Code Validation")
    L(f"Source: `GE-D-001-Department Abbreviations CODES.xlsx`")
    L("")
    if known_codes:
        L(f"**Known codes ({len(known_codes)}):** {', '.join(f'`{c}`' for c in sorted(known_codes))}")
    else:
        L("⚠️ Could not load GE-D-001 — dept code validation skipped")
    L("")
    if unknown_codes:
        L("**Unrecognised dept codes in filenames:**")
        for u in unknown_codes:
            L(f"- `{u['file']}` — code `{u['dept_code']}` not in GE-D-001")
    else:
        L("✅ All dept codes match GE-D-001")
    L("")

    # Data quality
    L("### Data Quality Issues")
    L("")

    mh = quality_issues['misspelled_headers']
    L(f"#### Misspelled Column Headers — {len(mh)} issue(s)")
    if mh:
        L("| File | Sheet | Found | Expected |")
        L("|------|-------|-------|----------|")
        for i in mh:
            L(f"| `{i['file']}` | `{i['sheet']}` | `{i['found']}` | `{i['expected']}` |")
    else:
        L("✅ No misspelled headers found")
    L("")

    di = quality_issues['disposition_issues']
    L(f"#### Disposition Value Issues (case / typos) — {len(di)} issue(s)")
    if di:
        L("| File | Sheet | Row | Value | Issue |")
        L("|------|-------|-----|-------|-------|")
        for i in di[:20]:
            L(f"| `{i['file']}` | `{i['sheet']}` | {i['row']} | `{i['value']}` | {i['issue']} |")
        if len(di) > 20:
            L(f"*… and {len(di) - 20} more rows*")
    else:
        L("✅ No disposition issues found")
    L("")

    rc = quality_issues['retention_case_issues']
    L(f"#### Retention Case Inconsistencies — {len(rc)} issue(s)")
    if rc:
        L("| File | Sheet | Row | Value | Issue |")
        L("|------|-------|-----|-------|-------|")
        for i in rc[:20]:
            L(f"| `{i['file']}` | `{i['sheet']}` | {i['row']} | `{i['value']}` | {i['issue']} |")
        if len(rc) > 20:
            L(f"*… and {len(rc) - 20} more rows*")
    else:
        L("✅ No retention case issues found")
    L("")

    rf = quality_issues['retention_contains_filename']
    L(f"#### Retention Fields Containing Filenames (data entry error) — {len(rf)} issue(s)")
    if rf:
        L("| File | Sheet | Row | Value |")
        L("|------|-------|-----|-------|")
        for i in rf:
            L(f"| `{i['file']}` | `{i['sheet']}` | {i['row']} | `{i['value']}` |")
    else:
        L("✅ No retention/filename mix-ups found")
    L("")

    # Summary
    total = (len(mh) + len(di) + len(rc) + len(rf)
             + len(invalid_files) + len(unknown_codes))
    L("---")
    L("## Summary")
    L("")
    L("| Category | Issues |")
    L("|----------|--------|")
    L(f"| Non-conforming filenames | {len(invalid_files)} |")
    L(f"| Unknown dept codes | {len(unknown_codes)} |")
    L(f"| Misspelled headers | {len(mh)} |")
    L(f"| Disposition value issues | {len(di)} |")
    L(f"| Retention case issues | {len(rc)} |")
    L(f"| Retention/filename mix-ups | {len(rf)} |")
    L(f"| **TOTAL** | **{total}** |")
    L("")

    return "\n".join(lines)


def save_report(md_content, output_dir):
    """Write the markdown report to a versioned file and print the saved path."""
    report_path = get_report_path(output_dir)
    report_path.write_text(md_content, encoding="utf-8")
    print(f"\n✅ Report saved → {report_path}")
    return report_path


def main():
    folder_path = r"C:\Users\bichn\MSBA\IMPACT\CoW-Data Lake\20260401-Compliance Report\01-Presales\Provided by CoW\Data Management Lists"

    report_output_dir = r"C:\Users\bichn\MSBA\IMPACT\CoW-Data Lake\Gap Analysis Report"

    run_ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    print("Analyzing Excel files for governance information...")
    data = deep_analyze_excel_files(folder_path)
    print("Analysis complete.\n")

    print_governance_report(data)

    # --- Validation against naming conventions & data quality ---
    print("\nRunning validation against naming conventions and data quality rules...")
    valid_files, invalid_files = validate_file_naming_conventions(folder_path)
    known_codes, unknown_codes = validate_dept_codes(folder_path, valid_files)
    quality_issues = validate_data_quality(folder_path)
    print_validation_report(valid_files, invalid_files, quality_issues, known_codes, unknown_codes)

    # --- Auto-save markdown report ---
    md = build_md_report(
        data, valid_files, invalid_files,
        quality_issues, known_codes, unknown_codes, run_ts
    )
    save_report(md, report_output_dir)

if __name__ == "__main__":
    main()
